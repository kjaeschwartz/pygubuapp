import os
import time
import tkinter as tk
import tkinter.ttk as ttk
import webbrowser
from tkinter.scrolledtext import ScrolledText
from datetime import date
import datetime
import shutil
import openpyxl
import pyautogui
#start py running pygubu-designer.exe
import pygubu

PROJECT_PATH = os.path.abspath(os.path.dirname(__file__))
PROJECT_UI = os.path.join(PROJECT_PATH, "ID_exporterVer2.0.ui")
id_list = []
globalColumnstartEnd = []
#TODO, FIND A WAY TO MAKE THE MESSAGE DYNAMIC.

class IdExporterver20App:
    def __init__(self, master=None):
        # build ui
        self.mainFrame = ttk.Frame(master)
        self.notebook = ttk.Notebook(self.mainFrame)
        self.tab1frame = ttk.Frame(self.notebook)
        self.exporterLabel = ttk.Label(self.tab1frame)
        self.mainLabel = tk.IntVar(value='ID Exporter')
        self.exporterLabel.configure(anchor='ne', borderwidth='2', compound='top', font='{Arial CYR} 14 {bold}')
        self.exporterLabel.configure(foreground='#276a70', relief='flat', state='disabled', style='Toolbutton')
        self.exporterLabel.configure(takefocus=True, text='ID Exporter', textvariable=self.mainLabel)
        self.exporterLabel.grid(column='0', row='0')
        self.LeftFrame2 = ttk.Frame(self.tab1frame)
        self.ID_entry = ScrolledText(self.LeftFrame2)
        self.ID_entry.configure(autoseparators='true', background='#f7fcfd', blockcursor='true', borderwidth='1')
        self.ID_entry.configure(height='12', highlightbackground='#69c4cb', highlightthickness='1', setgrid='false')
        self.ID_entry.configure(state='normal', tabstyle='wordprocessor', takefocus=False, undo='true')
        self.ID_entry.configure(width='20')
        self.ID_entry.pack(side='top')
        self.enterButton = tk.Button(self.LeftFrame2)
        self.enterButton.configure(background='#8ceaea', justify='left', relief='raised', text='Enter',command=self.retrieve_input)
        self.enterButton.pack(ipadx='13', padx='10', pady='10', side='top')
        self.LeftFrame2.configure(height='200', padding='10', relief='flat', width='200')
        self.LeftFrame2.grid(column='0', row='1')
        self.rightFrame = tk.Frame(self.tab1frame)
        self.inputATMbutton = tk.Button(self.rightFrame)
        self.inputATMbutton.configure(background='#8ceaea', foreground='#030a07', justify='left', padx='22',command=self.open_ids_ATM)
        self.inputATMbutton.configure(relief='raised', text='Input to ATM')
        self.inputATMbutton.grid(column='0', pady='7', row='0')
        self.sendGSTDBsheetButton = tk.Button(self.rightFrame)
        self.sendGSTDBsheetButton.configure(background='#8ceaea', compound='top', text='Send to GSTDB sheet',command=self.send2GSTDB_func)
        self.sendGSTDBsheetButton.grid(column='0', pady='7', row='1')
        self.makeGSTDBfoldersButton = tk.Button(self.rightFrame)
        self.makeGSTDBfoldersButton.configure(background='#8ceaea', justify='left', text='Make GSTDB folders',command=self.make_GSTDB_folders_func)
        self.makeGSTDBfoldersButton.grid(column='0', pady='7', row='2')
        self.CTsearchSetupButton = tk.Button(self.rightFrame)
        self.CTsearchSetupButton.configure(background='#8ceaea', cursor='arrow', justify='left', padx='13')
        self.CTsearchSetupButton.configure(relief='raised', text='CT search setup')
        self.CTsearchSetupButton.grid(column='0', pady='7', row='3')
        self.rightFrame.configure(height='170', highlightbackground='#bcb5e6',
                                  highlightcolor='#a8bef2')
        self.rightFrame.configure(padx='20', pady='20', takefocus=False, width='200')
        self.rightFrame.grid(column='2', row='1', sticky='n')
        self.tab1frame.columnconfigure('2', pad='0')
        self.tab1frame.configure(borderwidth='2', height='400', relief='flat', takefocus=True)
        self.tab1frame.configure(width='600')
        self.tab1frame.pack()
        self.notebook.add(self.tab1frame, compound='center', state='normal', sticky='nw', text='Main')
        self.tab2frame= ttk.Frame(self.notebook)
        self.Buttonframe2 = tk.Frame(self.tab2frame, container='false')
        self.send2FM_sheet = tk.Button(self.Buttonframe2)
        self.send2FM_sheet.configure(background='#8ceaea', foreground='#030a07', justify='left', padx='5')
        self.send2FM_sheet.configure(relief='raised', text='Send to FM sheet', width='10')
        self.send2FM_sheet.pack(expand='true', fill='x', pady='15', side='top')
        self.send2followup = tk.Button(self.Buttonframe2)
        self.send2followup.configure(background='#8ceaea', cursor='arrow', relief='raised',
                                     text='Send to follow up sheet')
        self.send2followup.pack(fill='x', pady='15', side='top')
        self.Send2IDarchive = tk.Button(self.Buttonframe2)
        self.Send2IDarchive.configure(background='#8ceaea', justify='center', takefocus=False, text='Send to ID archive',command=self.send2pendingArchiveFunc)
        self.Send2IDarchive.pack(fill='x', pady='15', side='top')
        self.sendNames2Archive = tk.Button(self.Buttonframe2)
        self.sendNames2Archive.configure(background='#8ceaea', cursor='arrow', justify='center', overrelief='raised')
        self.sendNames2Archive.configure(padx='20', relief='raised', state='normal', text='Send names to Archive',command=self.sendnames2pendingArchiveFunc)
        self.sendNames2Archive.configure(width='10')
        self.sendNames2Archive.pack(fill='x', pady='15', side='top')
        self.Buttonframe2.configure(height='170', padx='20', pady='20', takefocus=False)
        self.Buttonframe2.configure(width='100')
        self.Buttonframe2.pack(anchor='w', side='right')
        self.LeftFrame2 = tk.Frame(self.tab2frame, container='false')
        self.tab2label = ttk.Label(self.LeftFrame2)
        self.tab2label.configure(anchor='ne', borderwidth='2', compound='top', font='{Arial CYR} 14 {bold}')
        self.tab2label.configure(foreground='#276a70', relief='flat', state='disabled', style='Toolbutton')
        self.tab2label.configure(takefocus=True, text='Send to', textvariable=self.mainLabel)
        self.tab2label.pack()
        self.message2 = tk.Message(self.LeftFrame2)
        self.message2.configure(anchor='center', highlightbackground='#00c6c6', highlightcolor='#00c6c6',
                                highlightthickness='2')
        self.message2.configure(justify='center', relief='sunken', takefocus=False, text=id_list)
        self.message2.configure(width='200')
        self.message2.pack(expand='true', fill='both', padx='10', pady='10', side='top')
        self.LeftFrame2.configure(borderwidth='1',height='200', padx='30', pady='20',highlightbackground='#00ffff', highlightcolor='#00ffff',relief = 'sunken')
        self.LeftFrame2.configure(takefocus=False, width='180')
        self.LeftFrame2.pack(fill='both', side='left')
        self.LeftFrame2.pack_propagate(0)
        self.tab2frame.configure(height='400', relief='sunken', width='600')
        self.tab2frame.pack(side='right')
        self.tab2frame.pack_propagate(0)
        self.notebook.add(self.tab2frame, compound='top', sticky='n', text='Send To')
        self.templateFrame = ttk.Frame(self.notebook)
        self.subsection3 = tk.Frame(self.templateFrame, container='false')
        self.subsection3.configure(background='#dadada', height='170', highlightbackground='#bcb5e6',
                                   highlightcolor='#a8bef2')
        self.subsection3.configure(padx='20', pady='20', takefocus=False, width='100')
        self.subsection3.pack(anchor='e', side='right')
        self.templateFrame.configure(height='400', relief='sunken', width='600')
        self.templateFrame.pack(side='right')
        self.templateFrame.pack_propagate(0)
        self.notebook.add(self.templateFrame, text='Tools')
        self.templateframe = ttk.Frame(self.notebook)
        self.subsection4 = tk.Frame(self.templateframe, container='false')
        self.subsection4.configure(background='#dadada', height='170', highlightbackground='#bcb5e6',
                                   highlightcolor='#a8bef2')
        self.subsection4.configure(padx='20', pady='20', takefocus=False, width='100')
        self.subsection4.pack(anchor='e', side='right')
        self.templateframe.configure(height='400', relief='sunken', width='600')
        self.templateframe.pack(side='right')
        self.templateframe.pack_propagate(0)
        self.notebook.add(self.templateframe, text='ETA dates')
        self.notebook.configure(height='290', width='400')
        self.notebook.pack(side='top')
        self.mainFrame.configure(height='200', width='200')
        self.mainFrame.pack(side='top')

        # Main widget
        self.mainwindow = self.mainFrame

    def retrieve_input(self):
        id = self.ID_entry.get("1.0", 'end-1c')
        idListprototype = id.splitlines()
        # print(idListprototype)
        if len(id_list) != 0:
            id_list.clear()
            # print(f"id_list =={id_list}")
        for j in idListprototype:
            id_list.append(j)
        print(id_list)
        return id_list

    class open_ids():
        def make_chrome_window2(self):
            fw = pyautogui.getWindowsWithTitle('ATM - Google Chrome')
            pyautogui.scroll(200)
            if len(fw) == 0:
                print("l is 0")
                webbrowser.open('https://atm.accuratebackground.com/atm/login.jsp')
                fw = pyautogui.getWindowsWithTitle('Vendor Login | Accurate Background - Google Chrome')
            fw = fw[0]
            fw.width = 974
            fw.topleft = (953, 0)

        def get_new_tab(self):
            webbrowser.open('https://atm.accuratebackground.com/atm/findSearch.html')
            time.sleep(1.1)

        def search_id_fetch(self, ids_list):
            self.make_chrome_window2()
            # for h in range(0,len(ids_list)):
            for h in id_list:
                time.sleep(.05)
                self.get_new_tab()
                # findsearch = ["enter_id_box.png", "search_press_box.png"]
                time.sleep(.05)
                enter_id_box = (1340, 405)
                press_search_button = (1625, 405)
                pyautogui.click(enter_id_box)
                time.sleep(.05)
                pyautogui.typewrite(h)
                time.sleep(.15)
                pyautogui.click(press_search_button)

        # auto_start()
    def open_ids_ATM(self):
        ids_list = self.retrieve_input()
        self.open_ids().search_id_fetch(ids_list)
        pyautogui.hotkey('ctrl', '2', interval=.07)
        # Testimport = "test succeeded"
        #
    def make_GSTDB_folders_func(self):
        exec(open('C:\\Users\kschwartz\PycharmProjects\pythonProject\make_GSCDB_folders.py').read())
    def send2pendingArchiveFunc(self):
        if len(id_list) != 0:
            pendingIDspath = 'C:\\Users\kschwartz\Documents\CT_pending_IDS.xlsx'
            pendIDsWB = openpyxl.load_workbook(pendingIDspath)
            main_sheet= pendIDsWB['main_pending_ids']
            idcolumnList = main_sheet['B2':'B1001']
            def count_filled_entries():
                i = 0
                for rowOfCellObjects in idcolumnList:
                    for cellObj in rowOfCellObjects:
                        if cellObj.value != None:
                            i +=1
                return i
            def addNewEntries():
                filledCells = count_filled_entries()
                start_point = filledCells + 1
                globalColumnstartEnd.append(str(start_point))
                end_point = str(len(id_list) + filledCells)
                globalColumnstartEnd.append(str(end_point))
                print(f"GlobalColumnsStartEnd is {globalColumnstartEnd}")
                print(f"start point is {start_point}")
                print(f"end point is {end_point}")
                columnStart = "B" + str(start_point)
                columnEnd = "B" + str(end_point)
                print(f"Column start is {columnStart}")
                print(f"Column end is {columnEnd}")
                filledIDcolumn = main_sheet[columnStart:columnEnd]
                for rowOfCellObjects in filledIDcolumn:
                    for cellObj in rowOfCellObjects:
                        # print(cellObj.coordinate, cellObj.value)
                        cellObj.value = str(id_list[filledIDcolumn.index(rowOfCellObjects)])
                pendIDsWB.save("C:\\Users\kschwartz\Documents\CT_pending_IDS.xls")
                pendIDsWB.close()
                os.startfile("C:\\Users\kschwartz\Documents\CT_pending_IDS.xls")
            addNewEntries()
            print("sent to CT_pending_IDS workbook")
    def sendnames2pendingArchiveFunc(self):
        pendingIDspath = 'C:\\Users\kschwartz\Documents\CT_pending_IDS.xlsx'
        pendIDsWB = openpyxl.load_workbook(pendingIDspath)
        nameColStart = "C" + (globalColumnstartEnd[0] +1)
        nameColEnd = "C" + globalColumnstartEnd[1]
        main_sheet = pendIDsWB['main_pending_ids']
        nameColList = main_sheet[nameColStart:nameColEnd]
        for rowOfCellObjects in nameColList:
            for cellObj in rowOfCellObjects:
                # print(cellObj.coordinate, cellObj.value)
                cellObj.value = str(id_list[nameColList.index(rowOfCellObjects)])
        pendIDsWB.save("C:\\Users\kschwartz\Documents\CT_pending_IDS.xls")
        pendIDsWB.close()
        os.startfile("C:\\Users\kschwartz\Documents\CT_pending_IDS.xls")
        print("sent names to pendIDsWB")
    def send2GSTDB_func(self):
        if len(id_list) != 0:
            GSTDBsheetPath = "C:\\Users\kschwartz\Documents\GA-SCDB-Search-helper_realTEST.xlsm"
            GA_wb = openpyxl.load_workbook(GSTDBsheetPath)
            GA_sheet = GA_wb['main_sheet']

            def wipePreviousEntries():
                for rowOfCellObjects in GA_sheet['B2':'B51']:
                    for cellObj in rowOfCellObjects:
                        if cellObj.value != None:
                            cellObj.value = ''
                for rowOfCellObjects in GA_sheet['D2':'D51']:
                    for cellObj in rowOfCellObjects:
                        if cellObj.value != None:
                            cellObj.value = ''
                for rowOfCellObjects in GA_sheet['F2':'F51']:
                    for cellObj in rowOfCellObjects:
                        if cellObj.value != None:
                            cellObj.value = ''
                for rowOfCellObjects in GA_sheet['J2':'J12']:
                    for cellObj in rowOfCellObjects:
                        if cellObj.value != None:
                            cellObj.value = ''
                for rowOfCellObjects in GA_sheet['L2':'L12']:
                    for cellObj in rowOfCellObjects:
                        if cellObj.value != None:
                            cellObj.value = ''

            def addNewEntries():
                end_point = str(len(id_list) + 1)
                print(f"end point is {end_point}")
                columnEnd = "B" + end_point
                print(f"Column end is {columnEnd}")
                sheet2list = GA_sheet['B2':columnEnd]
                for rowOfCellObjects in sheet2list:
                    for cellObj in rowOfCellObjects:
                        # print(cellObj.coordinate, cellObj.value)
                        cellObj.value = str(id_list[sheet2list.index(rowOfCellObjects)])

            wipePreviousEntries()
            addNewEntries()
            GA_wb.save("C:\\Users\kschwartz\Documents\GA-SCDB-Search-helper_realTEST.xls")
            GA_wb.close()
            os.startfile("C:\\Users\kschwartz\Documents\GA-SCDB-Search-helper_realTEST.xls")
            return
    def sendIDS2junkWorkbook(self):
        genjunkWBpath = 'C:\\Users\kschwartz\Documents\CT_pending_IDS.xlsx'
        pendIDsWB = openpyxl.load_workbook(pendingIDspath)
        nameColStart = "C" + (globalColumnstartEnd[0] +1)
        nameColEnd = "C" + globalColumnstartEnd[1]
        main_sheet = pendIDsWB['main_pending_ids']
        # def cleartable():
        #     for a in sheet['A1':'A2']

    def run(self):
        self.mainwindow.mainloop()


if __name__ == '__main__':
    root = tk.Tk()
    app = IdExporterver20App(root)
    app.run()


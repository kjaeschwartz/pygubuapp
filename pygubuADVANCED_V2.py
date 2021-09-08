import os
import time
import tkinter as tk
import tkinter.ttk as ttk
import webbrowser
from tkinter.scrolledtext import ScrolledText
from datetime import date
import datetime
import shutil
import openpyxl
import pyautogui
#start py running pygubu-designer.exe
import pygubu


PROJECT_PATH = os.path.abspath(os.path.dirname(__file__))
PROJECT_UI = os.path.join(PROJECT_PATH, "uiADVANCEDBACKUP.ui")
id_list = []
globalColumnstartEnd = []
#TODO, FIND A WAY TO MAKE THE MESSAGE DYNAMIC.

class UiadvancedbackupApp:
    def __init__(self, master=None):
        # build ui
        self.mainFrame = ttk.Frame(master)
        self.notebook4 = ttk.Notebook(self.mainFrame)
        self.tab1frame = ttk.Frame(self.notebook4)
        self.exporterLabel = ttk.Label(self.tab1frame)
        self.mainLabel = tk.IntVar(value='ID Exporter')
        self.exporterLabel.configure(anchor='ne', borderwidth='2', compound='top', font='{Arial CYR} 14 {bold}')
        self.exporterLabel.configure(foreground='#276a70', relief='flat', state='disabled', style='Toolbutton')
        self.exporterLabel.configure(takefocus=True, text='ID Exporter', textvariable=self.mainLabel)
        self.exporterLabel.grid(column='0', row='0')
        self.leftFrame = ttk.Frame(self.tab1frame)
        self.ID_entry = ScrolledText(self.leftFrame)
        self.ID_entry.configure(autoseparators='true', background='#f7fcfd', blockcursor='true', borderwidth='1')
        self.ID_entry.configure(height='12', highlightbackground='#69c4cb', highlightthickness='1', setgrid='false')
        self.ID_entry.configure(state='normal', tabstyle='wordprocessor', takefocus=False, undo='true')
        self.ID_entry.configure(width='20')
        self.ID_entry.pack(side='top')
        self.enterButton = tk.Button(self.leftFrame)
        self.enterButton.configure(background='#8ceaea', justify='left', relief='raised', state='normal')
        self.enterButton.configure(text='Enter',command=self.retrieve_input)
        self.enterButton.pack(expand='true', fill='x', ipadx='13', padx='55', pady='10', side='top')
        self.leftFrame.configure(height='200', padding='10', relief='flat', width='200')
        self.leftFrame.grid(column='0', row='1')
        self.rightFrame = tk.Frame(self.tab1frame)
        self.inputATMbutton = tk.Button(self.rightFrame)
        self.inputATMbutton.configure(background='#8ceaea', compound='top', foreground='#030a07', justify='left')
        self.inputATMbutton.configure(padx='22', relief='raised', text='Input to ATM',command=self.open_ids_ATM)
        self.inputATMbutton.pack(expand='true', fill='x', pady='15', side='top')
        self.sendGSTDBsheetButton = tk.Button(self.rightFrame)
        self.sendGSTDBsheetButton.configure(background='#8ceaea', compound='top', text='Send to GSTDB sheet',command=self.send2GSTDB_func)
        self.sendGSTDBsheetButton.pack(anchor='s', expand='true', fill='x', pady='15', side='top')
        self.makeGSTDBfoldersButton = tk.Button(self.rightFrame)
        self.makeGSTDBfoldersButton.configure(background='#8ceaea', justify='left', text='Make GSTDB folders',command=self.make_GSTDB_folders_func)
        self.makeGSTDBfoldersButton.pack(fill='x', pady='15', side='top')
        self.CTsearchSetupButton = tk.Button(self.rightFrame)
        self.CTsearchSetupButton.configure(background='#8ceaea', cursor='arrow', justify='left', padx='13')
        self.CTsearchSetupButton.configure(relief='raised', text='CT search setup')
        self.CTsearchSetupButton.pack(fill='x', pady='15', side='top')
        self.rightFrame.configure(height='300', highlightbackground='#bcb5e6',
                                  highlightcolor='#a8bef2')
        self.rightFrame.configure(padx='20', takefocus=False, width='200')
        self.rightFrame.grid(column='1', row='1', sticky='n')
        self.rightFrame.grid_propagate(0)
        self.tab1frame.configure(borderwidth='2', height='400', padding='14 0', relief='flat', takefocus=True)
        self.tab1frame.configure(width='600')
        self.tab1frame.pack()
        self.notebook4.add(self.tab1frame, compound='center', state='normal', sticky='nw', text='Main')
        self.tab2frame = ttk.Frame(self.notebook4)
        self.tab2label = ttk.Label(self.tab2frame)
        self.tab2label.configure(borderwidth='2', compound='top', font='{Arial CYR} 14 {bold}', foreground='#276a70')
        self.tab2label.configure(justify='left', relief='flat', state='disabled', takefocus=True)
        self.tab2label.configure(text='Send to')
        self.tab2label.pack(anchor='nw', padx='50', side='top')
        self.Buttonframe2 = tk.Frame(self.tab2frame, container='false')
        self.send2FM_sheet = tk.Button(self.Buttonframe2)
        self.send2FM_sheet.configure(background='#8ceaea', foreground='#030a07', justify='left', padx='5')
        self.send2FM_sheet.configure(relief='raised', text='Send to FM sheet', width='10')
        self.send2FM_sheet.pack(expand='true', fill='x', pady='15', side='top')
        self.send2followup = tk.Button(self.Buttonframe2)
        self.send2followup.configure(background='#8ceaea', cursor='arrow', relief='raised',
                                     text='Send to follow up sheet')
        self.send2followup.pack(fill='x', pady='15', side='top')
        self.Send2IDarchive = tk.Button(self.Buttonframe2)
        self.Send2IDarchive.configure(background='#8ceaea', justify='center', takefocus=False, text='Send to ID archive')
        self.Send2IDarchive.pack(fill='x', pady='15', side='top')
        self.sendNames2Archive = tk.Button(self.Buttonframe2)
        self.sendNames2Archive.configure(background='#8ceaea', cursor='arrow', justify='center', overrelief='raised')
        self.sendNames2Archive.configure(padx='20', relief='raised', state='normal', text='Send to GSTDB sheet')
        self.sendNames2Archive.configure(width='10')
        self.sendNames2Archive.pack(fill='x', pady='15', side='top')
        self.Buttonframe2.configure(height='170', highlightbackground='#bcb5e6', highlightcolor='#a8bef2', padx='20')
        self.Buttonframe2.configure(pady='20', takefocus=False, width='100')
        self.Buttonframe2.pack(anchor='w', side='right')
        self.LeftFrame2 = tk.Frame(self.tab2frame)
        self.message1 = tk.Message(self.LeftFrame2)
        self.message1.configure(anchor='center', highlightbackground='#00c6c6', highlightcolor='#00c6c6',
                                highlightthickness='2')
        self.message1.configure(relief='flat', takefocus=False, text='id_list', width='200')
        self.message1.pack(expand='true', fill='both', side='top')
        self.LeftFrame2.configure(borderwidth='1', height='200', highlightbackground='#00ffff',
                                  highlightcolor='#00ffff')
        self.LeftFrame2.configure(padx='30', pady='20', relief='flat', takefocus=False)
        self.LeftFrame2.configure(width='180')
        self.LeftFrame2.pack(fill='both', side='left')
        self.LeftFrame2.pack_propagate(0)
        self.tab2frame.configure(height='400', relief='flat', width='600')
        self.tab2frame.pack(side='top')
        self.notebook4.add(self.tab2frame, compound='top', sticky='n', text='Send To')
        self.tab3frame = ttk.Frame(self.notebook4)
        self.LeftFrame3 = tk.Frame(self.tab3frame, container='false')
        self.LeftFrame3.configure(background='#dadada', borderwidth='1', highlightbackground='#bcb5e6',
                                  highlightcolor='#a8bef2')
        self.LeftFrame3.configure(padx='20', pady='20', relief='groove', takefocus=False)
        self.LeftFrame3.configure(width='200')
        self.LeftFrame3.pack(anchor='e', expand='false', fill='both', side='left')
        self.RightFrame3 = tk.Frame(self.tab3frame, container='false')
        self.RightFrame3.configure(background='#dadada', highlightbackground='#bcb5e6', highlightcolor='#a8bef2',
                                   padx='20')
        self.RightFrame3.configure(pady='20', takefocus=False, width='200')
        self.RightFrame3.pack(anchor='e', expand='false', fill='both', side='left')
        self.tab3frame.configure(height='400', relief='sunken', width='600')
        self.tab3frame.pack(side='right')
        self.tab3frame.pack_propagate(0)
        self.notebook4.add(self.tab3frame, text='Tools')
        self.tab4frame = ttk.Frame(self.notebook4)
        self.LeftFrame4 = tk.Frame(self.tab4frame, container='false')
        self.frame5 = tk.Frame(self.LeftFrame4)
        self.label3 = tk.Label(self.frame5)
        self.label3.configure(font='{Bahnschrift SemiCondensed} 14 {bold}', foreground='#276a70',
                              highlightbackground='#4d7356', highlightcolor='#276a70')
        self.label3.configure(highlightthickness='3', text='GASTDB ETA')
        self.label3.pack(anchor='s', fill='x', padx='1', pady='1', side='top')
        self.frame5.configure(background='#437c77', height='7', relief='groove', width='200')
        self.frame5.pack(fill='x', side='top')
        self.GASTDB_ETA_text = tk.Text(self.LeftFrame4)
        self.GASTDB_ETA_text.configure(font='{Calibri} 30 {}', height='1', insertborderwidth='0', width='5')
        _text_ = '''ETA'''
        self.GASTDB_ETA_text.insert('0.0', _text_)
        self.GASTDB_ETA_text.pack(anchor='center', expand='true', fill='x', side='top')
        self.button1 = ttk.Button(self.LeftFrame4)
        self.button1.configure(default='normal', text='Get ETA')
        self.button1.pack(anchor='n', expand='true', fill='x', padx='40', side='top')
        self.button1.bind('<1>', self.callback, add='+')
        self.LeftFrame4.configure(background='#dadada', borderwidth='1', highlightbackground='#bcb5e6',
                                  highlightcolor='#a8bef2')
        self.LeftFrame4.configure(padx='20', pady='20', relief='groove', takefocus=False)
        self.LeftFrame4.configure(width='200')
        self.LeftFrame4.pack(anchor='center', expand='true', fill='both', side='left')
        self.RightFrame4 = tk.Frame(self.tab4frame, container='false')
        self.frame6 = tk.Frame(self.RightFrame4)
        self.label5 = tk.Label(self.frame6)
        self.label5.configure(font='{Bahnschrift SemiCondensed} 14 {bold}', foreground='#276a70',
                              highlightbackground='#4d7356', highlightcolor='#276a70')
        self.label5.configure(highlightthickness='3', text='CT ETA')
        self.label5.pack(anchor='s', fill='x', padx='1', pady='1', side='top')
        self.frame6.configure(background='#437c77', height='7', relief='groove', width='200')
        self.frame6.pack(fill='x', side='top')
        self.text4 = tk.Text(self.RightFrame4)
        self.text4.configure(font='{Calibri} 30 {}', height='1', insertborderwidth='0', width='5')
        _text_ = '''ETA'''
        self.text4.insert('0.0', _text_)
        self.text4.pack(anchor='center', expand='true', fill='x', side='top')
        self.button3 = ttk.Button(self.RightFrame4)
        self.button3.configure(default='active', text='Get ETA')
        self.button3.pack(anchor='n', expand='true', fill='x', padx='40', side='top')
        self.button3.bind('<1>', self.callback, add='+')
        self.RightFrame4.configure(background='#dadada', highlightbackground='#bcb5e6', highlightcolor='#a8bef2',
                                   padx='20')
        self.RightFrame4.configure(pady='20', takefocus=False, width='200')
        self.RightFrame4.pack(anchor='center', expand='true', fill='both', padx='5', side='left')
        self.tab4frame.configure(height='400', relief='sunken', width='600')
        self.tab4frame.pack(side='top')
        self.tab4frame.pack_propagate(0)
        self.notebook4.add(self.tab4frame, text='ETA dates')
        self.frame2 = ttk.Frame(self.notebook4)
        self.leftFrame5 = tk.Frame(self.frame2, container='false')
        self.frame22 = tk.Frame(self.leftFrame5)
        self.DOBlabel = tk.Label(self.frame22)
        self.DOBlabel.configure(font='{Bahnschrift} 12 {bold}', foreground='#051212', relief='flat', text='DOB')
        self.DOBlabel.place(anchor='nw', rely='0.0', x='7', y='60')
        self.AgeLabel = tk.Label(self.frame22)
        self.AgeLabel.configure(font='{Bahnschrift} 12 {bold}', foreground='#030c0c', relief='flat', text='AGE')
        self.AgeLabel.place(anchor='nw', x='7', y='155')
        self.frame22.configure(height='280', width='40')
        self.frame22.grid(column='1', ipady='10', row='0')
        self.leftFrame5.rowconfigure('0', minsize='300')
        self.leftFrame5.columnconfigure('1', minsize='40')
        self.frame25 = tk.Frame(self.leftFrame5)
        self.DOBentrybox = tk.Entry(self.frame25)
        self.DOBentrybox.configure(font='{Arial} 24 {}', highlightbackground='#00ffff', highlightthickness='1',
                                   justify='left')
        self.DOBentrybox.configure(takefocus=False, width='160')
        _text_ = '''DOB'''
        self.DOBentrybox.delete('0', 'end')
        self.DOBentrybox.insert('0', _text_)
        self.DOBentrybox.place(anchor='center', relwidth='.8', x='80', y='65')
        self.AGEtext = tk.Entry(self.frame25)
        self.AGEtext.configure(font='{Arial} 24 {}')
        _text_ = '''AGE'''
        self.AGEtext.delete('0', 'end')
        self.AGEtext.insert('0', _text_)
        self.AGEtext.place(anchor='center', relwidth='0.8', x='80', y='160')
        self.getageButton = tk.Button(self.frame25)
        self.getageButton.configure(background='#78dcdc', text='Get Age')
        self.getageButton.place(anchor='nw', x='43', y='205')
        self.text9 = tk.Text(self.frame25)
        self.text9.configure(font='{ariel} 24 {}', height='1', highlightbackground='#00ffff', highlightcolor='#00ffff')
        self.text9.configure(highlightthickness='1', padx='9', width='6')
        _text_ = '''AGE'''
        self.text9.insert('0.0', _text_)
        self.text9.place(anchor='center', x='80', y='160')
        self.frame25.configure(height='280', relief='sunken', width='160')
        self.frame25.grid(column='2', ipady='10', row='0')
        self.leftFrame5.rowconfigure('0', minsize='300')
        self.leftFrame5.columnconfigure('2', minsize='160')
        self.label29 = tk.Label(self.leftFrame5)
        self.label29.configure(font='{Times New Roman} 12 {bold underline}', foreground='#2b6a80', justify='center',
                               text='AGE TOOL')
        self.label29.place(anchor='nw', height='20', width='200', x='0', y='3')
        self.canvas3 = tk.Canvas(self.leftFrame5)
        self.canvas3.configure(background='#2b6a80', borderwidth='1', closeenough='0', confine='false')
        self.canvas3.configure(height='40', relief='flat', takefocus=False, width='200')
        self.canvas3.place(anchor='n', height='6', width='400', y='20')
        self.leftFrame5.configure(background='#fcfefd', borderwidth='1', height='270', relief='sunken')
        self.leftFrame5.configure(width='200')
        self.leftFrame5.pack(side='left')
        self.rightFrame5 = tk.Frame(self.frame2, container='false')
        self.entry8 = tk.Entry(self.rightFrame5)
        self.entry8.configure(highlightbackground='#00ffff', highlightthickness='1')
        _text_ = '''LAST'''
        self.entry8.delete('0', 'end')
        self.entry8.insert('0', _text_)
        self.entry8.place(anchor='nw', bordermode='outside', relx='.3', rely='0.15', x='0', y='0')
        self.entry9 = tk.Entry(self.rightFrame5)
        self.entry9.configure(highlightbackground='#00ffff', highlightthickness='1')
        _text_ = '''FIRST'''
        self.entry9.delete('0', 'end')
        self.entry9.insert('0', _text_)
        self.entry9.place(anchor='nw', relx='0.3', rely='.3', x='0', y='0')
        self.entry11 = tk.Entry(self.rightFrame5)
        self.entry11.configure(highlightbackground='#00ffff', highlightthickness='1')
        _text_ = '''YOB'''
        self.entry11.delete('0', 'end')
        self.entry11.insert('0', _text_)
        self.entry11.place(anchor='nw', relx='.3', rely='.45', y='0')
        self.COMnameLabelFrame = tk.LabelFrame(self.rightFrame5)
        self.comNameLast = tk.Label(self.COMnameLabelFrame)
        self.comNameLast.configure(font='{Bahnschrift Condensed} 12 {}', justify='left', text='Last')
        self.comNameLast.place(anchor='nw', relx='.05', rely='0.01', x='1', y='0')
        self.ComNameFirst = tk.Label(self.COMnameLabelFrame)
        self.ComNameFirst.configure(font='{Bahnschrift Condensed} 12 {}', text='First')
        self.ComNameFirst.place(anchor='nw', relx='.05', rely='.37', x='1', y='0')
        self.ComNameYOB = tk.Label(self.COMnameLabelFrame)
        self.ComNameYOB.configure(font='{Bahnschrift Condensed} 12 {}', padx='4', text='YOB')
        self.ComNameYOB.place(anchor='nw', relx='.05', rely='.8', x='0', y='0')
        self.COMnameLabelFrame.configure(borderwidth='1', height='100', relief='flat', width='200')
        self.COMnameLabelFrame.place(anchor='nw', height='110', width='40', x='0', y='40')
        self.commnameANSWER = tk.Text(self.rightFrame5)
        self.commnameANSWER.configure(background='#e1f7d5', blockcursor='false', borderwidth='2',
                                      font='{Cambria} 14 {bold}')
        self.commnameANSWER.configure(foreground='#091801', height='2', insertofftime='0', padx='20')
        self.commnameANSWER.configure(pady='15', relief='sunken', selectborderwidth='0', setgrid='false')
        self.commnameANSWER.configure(tabs='1', tabstyle='tabular', takefocus=False, width='10')
        self.commnameANSWER.configure(wrap='word')
        _text_ = '''MEETS ID POLICY'''
        self.commnameANSWER.insert('0.0', _text_)
        self.commnameANSWER.place(anchor='nw', bordermode='outside', x='20', y='199')
        self.ComNameEnter = tk.Button(self.rightFrame5)
        self.ComNameEnter.configure(background='#00c1c1', foreground='#020202', text='ENTER')
        self.ComNameEnter.place(anchor='nw', x='70', y='160')
        self.frame27 = tk.Frame(self.rightFrame5, container='false')
        self.ComNameLabel = tk.Label(self.frame27)
        self.ComNameLabel.configure(font='{Times New Roman} 12 {bold underline}', foreground='#2b6a80',
                                    justify='center', text='COMMON NAME TOOL')
        self.ComNameLabel.pack(expand='true', fill='both', padx='5', side='top')
        self.frame27.configure(height='30', width='200')
        self.frame27.place(anchor='nw', x='0', y='0')
        self.rightFrame5.configure(borderwidth='2', height='200', highlightbackground='#2b6a80',
                                   highlightcolor='#2b6a80')
        self.rightFrame5.configure(highlightthickness='1', relief='raised', width='200')
        self.rightFrame5.pack(expand='true', fill='both', side='top')
        self.frame2.configure(height='400', relief='sunken', width='400')
        self.frame2.pack(side='top')
        self.notebook4.add(self.frame2, text='Name/Age')
        self.notebook4.configure(height='290', width='400')
        self.notebook4.pack(side='top')
        self.mainFrame.configure(height='200', width='200')
        self.mainFrame.pack(side='top')

        # Main widget
        self.mainwindow = self.mainFrame
    def retrieve_input(self):
        id = self.ID_entry.get("1.0", 'end-1c')
        idListprototype = id.splitlines()
        # print(idListprototype)
        if len(id_list) != 0:
            id_list.clear()
            # print(f"id_list =={id_list}")
        for j in idListprototype:
            id_list.append(j)
        print(id_list)
        return id_list

    class open_ids():
        def make_chrome_window2(self):
            fw = pyautogui.getWindowsWithTitle('ATM - Google Chrome')
            pyautogui.scroll(200)
            if len(fw) == 0:
                print("l is 0")
                webbrowser.open('https://atm.accuratebackground.com/atm/login.jsp')
                fw = pyautogui.getWindowsWithTitle('Vendor Login | Accurate Background - Google Chrome')
            fw = fw[0]
            fw.width = 974
            fw.topleft = (953, 0)

        def get_new_tab(self):
            webbrowser.open('https://atm.accuratebackground.com/atm/findSearch.html')
            time.sleep(1.1)

        def search_id_fetch(self, ids_list):
            self.make_chrome_window2()
            # for h in range(0,len(ids_list)):
            for h in id_list:
                time.sleep(.05)
                self.get_new_tab()
                # findsearch = ["enter_id_box.png", "search_press_box.png"]
                time.sleep(.05)
                enter_id_box = (1340, 405)
                press_search_button = (1625, 405)
                pyautogui.click(enter_id_box)
                time.sleep(.05)
                pyautogui.typewrite(h)
                time.sleep(.15)
                pyautogui.click(press_search_button)

        # auto_start()
    def open_ids_ATM(self):
        ids_list = self.retrieve_input()
        self.open_ids().search_id_fetch(ids_list)
        pyautogui.hotkey('ctrl', '2', interval=.07)
        # Testimport = "test succeeded"
        #
    def make_GSTDB_folders_func(self):
        exec(open('C:\\Users\kschwartz\PycharmProjects\pythonProject\make_GSCDB_folders.py').read())
    def send2pendingArchiveFunc(self):
        if len(id_list) != 0:
            pendingIDspath = 'C:\\Users\kschwartz\Documents\CT_pending_IDS.xlsx'
            pendIDsWB = openpyxl.load_workbook(pendingIDspath)
            main_sheet= pendIDsWB['main_pending_ids']
            idcolumnList = main_sheet['B2':'B1001']
            def count_filled_entries():
                i = 0
                for rowOfCellObjects in idcolumnList:
                    for cellObj in rowOfCellObjects:
                        if cellObj.value != None:
                            i +=1
                return i
            def addNewEntries():
                filledCells = count_filled_entries()
                start_point = filledCells + 1
                globalColumnstartEnd.append(str(start_point))
                end_point = str(len(id_list) + filledCells)
                globalColumnstartEnd.append(str(end_point))
                print(f"GlobalColumnsStartEnd is {globalColumnstartEnd}")
                print(f"start point is {start_point}")
                print(f"end point is {end_point}")
                columnStart = "B" + str(start_point)
                columnEnd = "B" + str(end_point)
                print(f"Column start is {columnStart}")
                print(f"Column end is {columnEnd}")
                filledIDcolumn = main_sheet[columnStart:columnEnd]
                for rowOfCellObjects in filledIDcolumn:
                    for cellObj in rowOfCellObjects:
                        # print(cellObj.coordinate, cellObj.value)
                        cellObj.value = str(id_list[filledIDcolumn.index(rowOfCellObjects)])
                pendIDsWB.save("C:\\Users\kschwartz\Documents\CT_pending_IDS.xls")
                pendIDsWB.close()
                os.startfile("C:\\Users\kschwartz\Documents\CT_pending_IDS.xls")
            addNewEntries()
            print("sent to CT_pending_IDS workbook")
    def sendnames2pendingArchiveFunc(self):
        pendingIDspath = 'C:\\Users\kschwartz\Documents\CT_pending_IDS.xlsx'
        pendIDsWB = openpyxl.load_workbook(pendingIDspath)
        nameColStart = "C" + (globalColumnstartEnd[0] +1)
        nameColEnd = "C" + globalColumnstartEnd[1]
        main_sheet = pendIDsWB['main_pending_ids']
        nameColList = main_sheet[nameColStart:nameColEnd]
        for rowOfCellObjects in nameColList:
            for cellObj in rowOfCellObjects:
                # print(cellObj.coordinate, cellObj.value)
                cellObj.value = str(id_list[nameColList.index(rowOfCellObjects)])
        pendIDsWB.save("C:\\Users\kschwartz\Documents\CT_pending_IDS.xls")
        pendIDsWB.close()
        os.startfile("C:\\Users\kschwartz\Documents\CT_pending_IDS.xls")
        print("sent names to pendIDsWB")
    def send2GSTDB_func(self):
        if len(id_list) != 0:
            GSTDBsheetPath = "C:\\Users\kschwartz\Documents\GA-SCDB-Search-helper_realTEST.xlsm"
            GA_wb = openpyxl.load_workbook(GSTDBsheetPath)
            GA_sheet = GA_wb['main_sheet']

            def wipePreviousEntries():
                for rowOfCellObjects in GA_sheet['B2':'B51']:
                    for cellObj in rowOfCellObjects:
                        if cellObj.value != None:
                            cellObj.value = ''
                for rowOfCellObjects in GA_sheet['D2':'D51']:
                    for cellObj in rowOfCellObjects:
                        if cellObj.value != None:
                            cellObj.value = ''
                for rowOfCellObjects in GA_sheet['F2':'F51']:
                    for cellObj in rowOfCellObjects:
                        if cellObj.value != None:
                            cellObj.value = ''
                for rowOfCellObjects in GA_sheet['J2':'J12']:
                    for cellObj in rowOfCellObjects:
                        if cellObj.value != None:
                            cellObj.value = ''
                for rowOfCellObjects in GA_sheet['L2':'L12']:
                    for cellObj in rowOfCellObjects:
                        if cellObj.value != None:
                            cellObj.value = ''

            def addNewEntries():
                end_point = str(len(id_list) + 1)
                print(f"end point is {end_point}")
                columnEnd = "B" + end_point
                print(f"Column end is {columnEnd}")
                sheet2list = GA_sheet['B2':columnEnd]
                for rowOfCellObjects in sheet2list:
                    for cellObj in rowOfCellObjects:
                        # print(cellObj.coordinate, cellObj.value)
                        cellObj.value = str(id_list[sheet2list.index(rowOfCellObjects)])

            wipePreviousEntries()
            addNewEntries()
            GA_wb.save("C:\\Users\kschwartz\Documents\GA-SCDB-Search-helper_realTEST.xls")
            GA_wb.close()
            print("sent to sheet")
            os.startfile("C:\\Users\kschwartz\Documents\GA-SCDB-Search-helper_realTEST.xls")
            return
    def sendIDS2junkWorkbook(self):
        genjunkWBpath = 'C:\\Users\kschwartz\Documents\CT_pending_IDS.xlsx'
        pendIDsWB = openpyxl.load_workbook(pendingIDspath)
        nameColStart = "C" + (globalColumnstartEnd[0] +1)
        nameColEnd = "C" + globalColumnstartEnd[1]
        main_sheet = pendIDsWB['main_pending_ids']
        # def cleartable():
        #     for a in sheet['A1':'A2']

    def run(self):
        self.mainwindow.mainloop()

    def callback(self, event=None):
        pass

    def run(self):
        self.mainwindow.mainloop()


if __name__ == '__main__':
    root = tk.Tk()
    app = UiadvancedbackupApp(root)
    app.run()

